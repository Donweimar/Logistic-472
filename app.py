import streamlit as st
import pandas as pd
import plotly.express as px
import os
import io
from datetime import datetime

# Importaciones para generación de Excel estilizado (openpyxl)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Importaciones para generación de PDF profesional (ReportLab)
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# -----------------------------------------------------------------------------
# 1. Configuración General de la Aplicación
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title="Dashboard Logístico Nacional", 
    page_icon="🚚",
    layout="wide", 
    initial_sidebar_state="expanded"
)

# Estilos CSS adaptables a Modo Claro / Oscuro
st.markdown("""
    <style>
    .block-container { padding-top: 1.5rem; padding-bottom: 2rem; }
    
    div[data-testid="stMetric"] {
        background-color: var(--secondary-background-color);
        padding: 14px;
        border-radius: 10px;
        border: 1px solid rgba(128, 128, 128, 0.2);
    }
    
    div[data-testid="stMetricValue"], div[data-testid="stMetricLabel"] {
        color: var(--text-color) !important;
    }
    
    .report-card {
        background-color: var(--secondary-background-color);
        padding: 20px;
        border-radius: 10px;
        border: 1px solid rgba(128, 128, 128, 0.2);
        margin-bottom: 15px;
    }
    </style>
""", unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# 2. Función de Carga y Limpieza de Datos
# -----------------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def process_data(file_source):
    try:
        df = pd.read_csv(file_source, sep=";", encoding="latin1")
    except Exception:
        df = pd.read_csv(file_source, sep=",", encoding="utf-8", errors="ignore")
    
    df.columns = df.columns.str.strip()
    
    def clean_numeric(x):
        if pd.isna(x):
            return 0.0
        x_str = str(x).replace('%', '').strip()
        if not x_str or x_str in ['-', '--', ' - ', ' -', 'NaN', 'nan', 'None']:
            return 0.0
        try:
            if '.' in x_str and ',' not in x_str:
                parts = x_str.split('.')
                if len(parts[-1]) == 3:
                    x_str = x_str.replace('.', '')
            elif ',' in x_str:
                x_str = x_str.replace('.', '').replace(',', '.')
            return float(x_str)
        except (ValueError, TypeError):
            return 0.0

    cols_numericas = ['MANIFIESTOS', 'PIEZAS', 'PESO', 'VOLUME', 'CAPACIDAD']
    for col in cols_numericas:
        if col in df.columns:
            df[col] = df[col].apply(clean_numeric)
        else:
            df[col] = 0.0

    text_cols = ['AÑO', 'RUTA', 'OPERADOR', 'PLACA', 'HORA', 'NOMBRE_CONDUCTOR', 'CONDUCTOR']
    for text_col in text_cols:
        if text_col in df.columns:
            df[text_col] = df[text_col].astype(str).str.strip()
        else:
            if text_col == 'NOMBRE_CONDUCTOR' and 'CONDUCTOR' in df.columns:
                df['NOMBRE_CONDUCTOR'] = df['CONDUCTOR'].astype(str).str.strip()
            elif text_col == 'CONDUCTOR' and 'NOMBRE_CONDUCTOR' in df.columns:
                df['CONDUCTOR'] = df['NOMBRE_CONDUCTOR'].astype(str).str.strip()
            else:
                df[text_col] = "N/A"

    if 'HORA' in df.columns:
        df['HORA_SIMPLE'] = df['HORA'].str.split(':').str[0].str.zfill(2) + ":00"
    
    return df


# -----------------------------------------------------------------------------
# 3. Módulos de Generación de Reportes (Excel y PDF)
# -----------------------------------------------------------------------------
def generate_excel_report(df_filtered, df_annual):
    wb = Workbook()
    
    font_title = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="595959")
    font_sec_header = Font(name="Calibri", size=12, bold=True, color="1F4E78")
    font_tbl_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10)
    font_kpi_val = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    font_kpi_lbl = Font(name="Calibri", size=9, color="595959")
    
    fill_navy = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_kpi = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # --- HOJA 1: RESUMEN EJECUTIVO ---
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1['A1'] = "INFORME EJECUTIVO DE GESTIÓN LOGÍSTICA"
    ws1['A1'].font = font_title
    ws1['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws1['A2'].font = font_subtitle
    
    # Tarjetas KPIs
    kpis = [
        ("Total Viajes", len(df_filtered), "#,##0"),
        ("Manifiestos Totales", df_filtered['MANIFIESTOS'].sum(), "#,##0"),
        ("Peso Total (kg)", df_filtered['PESO'].sum(), "#,##0.0"),
        ("Piezas Totales", df_filtered['PIEZAS'].sum(), "#,##0"),
        ("Ocupación Promed.", df_filtered['VOLUME'].mean()/100, "0.0%")
    ]
    
    ws1['A4'] = "INDICADORES CLAVE DE RENDIMIENTO (KPIs)"
    ws1['A4'].font = font_sec_header
    
    cols = ['A', 'C', 'E', 'G', 'I']
    for idx, (label, val, fmt) in enumerate(kpis):
        col = cols[idx]
        col_next = chr(ord(col) + 1)
        
        ws1.merge_cells(f"{col}5:{col_next}5")
        ws1.merge_cells(f"{col}6:{col_next}6")
        
        c_lbl = ws1[f"{col}5"]
        c_lbl.value = label.upper()
        c_lbl.font = font_kpi_lbl
        c_lbl.alignment = Alignment(horizontal="center", vertical="center")
        
        c_val = ws1[f"{col}6"]
        c_val.value = val
        c_val.font = font_kpi_val
        c_val.number_format = fmt
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        
        for r in range(5, 7):
            for c in [col, col_next]:
                cell = ws1[f"{c}{r}"]
                cell.fill = fill_kpi
                cell.border = border_thin

    # Tabla Resumen Anual
    start_row = 9
    ws1.cell(row=start_row, column=1, value="RESUMEN ACUMULADO POR AÑO").font = font_sec_header
    
    headers_anual = ["Año", "Manifiestos", "Peso Total (kg)", "Piezas Totales", "Ocupación Promed. (%)"]
    for col_idx, h in enumerate(headers_anual, 1):
        cell = ws1.cell(row=start_row+1, column=col_idx, value=h)
        cell.font = font_tbl_header
        cell.fill = fill_navy
        cell.alignment = Alignment(horizontal="center")
    
    r_idx = start_row + 2
    for _, row in df_annual.iterrows():
        vals = [str(row['AÑO']), row['MANIFIESTOS'], row['PESO'], row['PIEZAS'], row['CAPACIDAD']]
        fmts = ["@", "#,##0", "#,##0.0", "#,##0", "0.0"]
        for c_idx, (v, f) in enumerate(zip(vals, fmts), 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=v)
            cell.font = font_data
            cell.number_format = f
            cell.border = border_thin
            if r_idx % 2 == 0:
                cell.fill = fill_zebra
        r_idx += 1

    # --- HOJA 2: DETALLE DE OPERACIÓN ---
    ws2 = wb.create_sheet(title="Detalle de Operación")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2['A1'] = "DETALLE COMPLETO DE REGISTROS DE VIAJES"
    ws2['A1'].font = font_title
    
    headers_det = list(df_filtered.columns)
    for col_idx, h in enumerate(headers_det, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = font_tbl_header
        cell.fill = fill_navy
        cell.alignment = Alignment(horizontal="center")
        
    for row_idx, row in df_filtered.iterrows():
        curr_row = row_idx + 4
        for col_idx, val in enumerate(row, 1):
            cell = ws2.cell(row=curr_row, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_thin
            if curr_row % 2 == 0:
                cell.fill = fill_zebra
            
            col_name = headers_det[col_idx-1]
            if col_name in ['MANIFIESTOS', 'PIEZAS']:
                cell.number_format = "#,##0"
            elif col_name in ['PESO', 'VOLUME', 'CAPACIDAD']:
                cell.number_format = "#,##0.0"

    # Ajuste automático de anchos de columna
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_pdf_report(df_filtered, df_annual):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=18, leading=22, textColor=colors.HexColor('#1F4E78'), spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle', parent=styles['Normal'], fontName='Helvetica', fontSize=9, leading=11, textColor=colors.HexColor('#595959'), spaceAfter=12
    )
    sec_heading = ParagraphStyle(
        'SecHeading', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=12, leading=15, textColor=colors.HexColor('#1F4E78'), spaceBefore=10, spaceAfter=6
    )
    body_style = ParagraphStyle(
        'BodyTextCustom', parent=styles['Normal'], fontName='Helvetica', fontSize=9, leading=12, textColor=colors.HexColor('#262626')
    )
    cell_header = ParagraphStyle(
        'CellHeader', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, leading=11, textColor=colors.white, alignment=1
    )
    cell_body = ParagraphStyle(
        'CellBody', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10, alignment=1
    )

    story = []
    
    # Encabezado principal
    story.append(Paragraph("INFORME EJECUTIVO DE GESTIÓN LOGÍSTICA", title_style))
    story.append(Paragraph(f"<b>Fecha de emisión:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')} | <b>Total Registros:</b> {len(df_filtered):,}", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1F4E78'), spaceAfter=12))
    
    # 1. Tabla KPIs
    story.append(Paragraph("1. Indicadores Clave de Desempeño (KPIs)", sec_heading))
    kpi_data = [
        [
            Paragraph("<b>Total Viajes</b>", cell_header),
            Paragraph("<b>Manifiestos</b>", cell_header),
            Paragraph("<b>Peso Total (kg)</b>", cell_header),
            Paragraph("<b>Piezas Totales</b>", cell_header),
            Paragraph("<b>Ocupación Prom.</b>", cell_header)
        ],
        [
            Paragraph(f"<b>{len(df_filtered):,}</b>", cell_body),
            Paragraph(f"<b>{int(df_filtered['MANIFIESTOS'].sum()):,}</b>", cell_body),
            Paragraph(f"<b>{df_filtered['PESO'].sum():,.0f}</b>", cell_body),
            Paragraph(f"<b>{int(df_filtered['PIEZAS'].sum()):,}</b>", cell_body),
            Paragraph(f"<b>{df_filtered['VOLUME'].mean():.1f}%</b>", cell_body)
        ]
    ]
    t_kpi = Table(kpi_data, colWidths=[105]*5)
    t_kpi.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')),
        ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#F2F4F7')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t_kpi)
    story.append(Spacer(1, 10))
    
    # 2. Resumen Anual
    story.append(Paragraph("2. Resumen de Flujo de Carga por Año", sec_heading))
    annual_data = [[
        Paragraph("<b>Año</b>", cell_header),
        Paragraph("<b>Manifiestos</b>", cell_header),
        Paragraph("<b>Peso Total (kg)</b>", cell_header),
        Paragraph("<b>Piezas Totales</b>", cell_header),
        Paragraph("<b>Ocupación Promed. (%)</b>", cell_header)
    ]]
    for _, row in df_annual.iterrows():
        annual_data.append([
            Paragraph(str(row['AÑO']), cell_body),
            Paragraph(f"{int(row['MANIFIESTOS']):,}", cell_body),
            Paragraph(f"{row['PESO']:,.0f}", cell_body),
            Paragraph(f"{int(row['PIEZAS']):,}", cell_body),
            Paragraph(f"{row['VOLUME']:.1f}%", cell_body)
        ])
    t_annual = Table(annual_data, colWidths=[105]*5)
    t_annual.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t_annual)
    story.append(Spacer(1, 10))

    # 3. Rankings
    story.append(Paragraph("3. Rankings Principales (Top Conductores y Rutas)", sec_heading))
    df_cond = df_filtered.groupby('NOMBRE_CONDUCTOR').size().reset_index(name='VIAJES')
    df_cond = df_cond[df_cond['NOMBRE_CONDUCTOR'] != 'N/A'].sort_values('VIAJES', ascending=False).head(5)
    
    df_rutas = df_filtered.groupby('RUTA')['PESO'].sum().reset_index()
    df_rutas = df_rutas.sort_values('PESO', ascending=False).head(5)
    
    cond_rows = [[Paragraph("<b>Conductor</b>", cell_header), Paragraph("<b>Viajes</b>", cell_header)]]
    for _, r in df_cond.iterrows():
        cond_rows.append([Paragraph(str(r['NOMBRE_CONDUCTOR'])[:20], cell_body), Paragraph(str(r['VIAJES']), cell_body)])
    t_cond = Table(cond_rows, colWidths=[170, 80])
    t_cond.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#3B82F6')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    
    ruta_rows = [[Paragraph("<b>Ruta</b>", cell_header), Paragraph("<b>Peso (kg)</b>", cell_header)]]
    for _, r in df_rutas.iterrows():
        ruta_rows.append([Paragraph(str(r['RUTA'])[:22], cell_body), Paragraph(f"{r['PESO']:,.0f}", cell_body)])
    t_rutas = Table(ruta_rows, colWidths=[170, 80])
    t_rutas.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#06B6D4')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    
    t_rankings = Table([[t_cond, t_rutas]], colWidths=[260, 260])
    t_rankings.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(t_rankings)
    story.append(Spacer(1, 12))

    # 4. Observaciones
    story.append(Paragraph("4. Observaciones y Hallazgos Operativos", sec_heading))
    sub_viajes = df_filtered[df_filtered['VOLUME'] < 30].shape[0]
    pct_sub = (sub_viajes / len(df_filtered)) * 100 if len(df_filtered) > 0 else 0
    
    obs_text = f"""
    • <b>Estructura de Capacidad:</b> Se identificaron <b>{sub_viajes} viajes ({pct_sub:.1f}%)</b> operando por debajo del 30% de capacidad.<br/>
    • <b>Promedio General de Ocupación:</b> La flota registra un promedio global de <b>{df_filtered['CAPACIDAD'].mean():.1f}%</b>.<br/>
    • <b>Generación Automatizada:</b> Documento compilado en tiempo real mediante el motor de analítica del Dashboard.
    """
    story.append(Paragraph(obs_text, body_style))

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


# -----------------------------------------------------------------------------
# 4. Barra Lateral (Sidebar): Logo, Datos y Filtros
# -----------------------------------------------------------------------------
st.sidebar.title("⚙️ Configuración")

# Logo corporativo embebido (Ruta fija)
LOGO_PATH = "Logo.png" if os.path.exists("Logo.png") else ("logo.png" if os.path.exists("logo.png") else None)

if LOGO_PATH:
    st.sidebar.image(LOGO_PATH, use_container_width=True)

st.sidebar.markdown("---")

# Carga de archivo CSV
uploaded_file = st.sidebar.file_uploader("Importar archivo CSV", type=["csv"])

if uploaded_file is not None:
    df_raw = process_data(uploaded_file)
    st.sidebar.success("✅ Datos cargados correctamente")
elif os.path.exists("RutaNacional.csv"):
    df_raw = process_data("RutaNacional.csv")
    st.sidebar.info("ℹ️ Usando 'RutaNacional.csv'")
else:
    st.warning("⚠️ Por favor sube un archivo `.csv` en la barra lateral.")
    st.stop()

# Filtros dinámicos
st.sidebar.subheader("Filtros de Operación")

anos_disponibles = sorted([a for a in df_raw['AÑO'].unique() if a not in ['nan', 'N/A']])
selected_anos = st.sidebar.multiselect("Año(s)", anos_disponibles)

operadores_disponibles = sorted([op for op in df_raw['OPERADOR'].unique() if op not in ['nan', 'N/A']])
selected_operadores = st.sidebar.multiselect("Operador(es) Logístico(s)", operadores_disponibles)

rutas_disponibles = sorted([r for r in df_raw['RUTA'].unique() if r not in ['nan', 'N/A']])
selected_rutas = st.sidebar.multiselect("Ruta(s)", rutas_disponibles)

conductores_disponibles = sorted([c for c in df_raw['NOMBRE_CONDUCTOR'].unique() if c not in ['nan', 'N/A']])
selected_conductores = st.sidebar.multiselect("Conductor(es)", conductores_disponibles)


# -----------------------------------------------------------------------------
# 5. Aplicación de Filtros
# -----------------------------------------------------------------------------
df_filtered = df_raw.copy()

if selected_anos:
    df_filtered = df_filtered[df_filtered['AÑO'].isin(selected_anos)]

if selected_operadores:
    df_filtered = df_filtered[df_filtered['OPERADOR'].isin(selected_operadores)]

if selected_rutas:
    df_filtered = df_filtered[df_filtered['RUTA'].isin(selected_rutas)]

if selected_conductores:
    df_filtered = df_filtered[df_filtered['NOMBRE_CONDUCTOR'].isin(selected_conductores)]

if df_filtered.empty:
    st.error("No se encontraron registros con los filtros seleccionados.")
    st.stop()


# -----------------------------------------------------------------------------
# 6. Encabezado e Indicadores Clave (KPIs)
# -----------------------------------------------------------------------------
header_col1, header_col2 = st.columns([4, 1])

with header_col1:
    st.title("🚚 Dashboard de Gestión Logística Nacional")
    st.markdown("Visualización estratégica del flujo operativo, capacidad y eficiencia.")

with header_col2:
    if LOGO_PATH:
        st.image(LOGO_PATH, width=150)

# Alertas automáticas
viajes_subutilizados = df_filtered[df_filtered['VOLUME'] < 30].shape[0]
total_viajes = df_filtered.shape[0]
pct_sub = (viajes_subutilizados / total_viajes) * 100 if total_viajes > 0 else 0

if pct_sub > 15:
    st.warning(f"⚠️ **Alerta de Capacidad:** El **{pct_sub:.1f}%** de los viajes registrados ({viajes_subutilizados} despachos) salieron con menos del 30% de ocupación.")

# Métricas KPIs
kpi1, kpi2, kpi3, kpi4, kpi5 = st.columns(5)
kpi1.metric("Viajes Realizados", f"{len(df_filtered):,}")
kpi2.metric("Envíos Totales (Manifiestos)", f"{int(df_filtered['MANIFIESTOS'].sum()):,}")
kpi3.metric("Peso Total Movilizado", f"{df_filtered['PESO'].sum():,.0f} kg")
kpi4.metric("Piezas Transportadas", f"{int(df_filtered['PIEZAS'].sum()):,}")
kpi5.metric("Ocupación Promedio", f"{df_filtered['VOLUME'].mean():.1f}%")

st.markdown("---")


# -----------------------------------------------------------------------------
# 7. Preparación de Agregación Anual
# -----------------------------------------------------------------------------
df_annual = df_filtered.groupby('AÑO').agg({
    'MANIFIESTOS': 'sum',
    'PESO': 'sum',
    'CAPACIDAD': 'sum',
    'PIEZAS': 'sum',
    'VOLUME': 'mean'
}).reset_index()


# -----------------------------------------------------------------------------
# 8. Organización en Pestañas (Tabs)
# -----------------------------------------------------------------------------
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 Métricas Anuales", 
    "👨‍✈️ Conductores y Rutas", 
    "⏱️ Operación y Eficiencia", 
    "📄 Generación de Reportes",
    "📋 Datos Auditables"
])

# --- TAB 1: MÉTRICAS ANUALES ---
with tab1:
    st.subheader("Evolución y Totales Anuales")
    
    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown("##### 📦 Total de PESO Movilizado por Año (kg)")
        fig_peso_ano = px.bar(
            df_annual, x='AÑO', y='PESO',
            text_auto='.2s', template='plotly_white',
            color_discrete_sequence=['#f59e0b']
        )
        fig_peso_ano.update_layout(xaxis_title="Año", yaxis_title="Peso Total (kg)", margin=dict(l=10, r=10, t=30, b=10))
        st.plotly_chart(fig_peso_ano, use_container_width=True)

    with col_b:
        st.markdown("##### 🧩 Total de PIEZAS Transportadas por Año")
        fig_piezas_ano = px.bar(
            df_annual, x='AÑO', y='PIEZAS',
            text_auto='.2s', template='plotly_white',
            color_discrete_sequence=['#10b981']
        )
        fig_piezas_ano.update_layout(xaxis_title="Año", yaxis_title="Piezas Totales", margin=dict(l=10, r=10, t=30, b=10))
        st.plotly_chart(fig_piezas_ano, use_container_width=True)

    col_c, col_d = st.columns(2)
    with col_c:
        st.markdown("##### 📄 Envíos Totales (Manifiestos por Año)")
        fig_envios = px.bar(
            df_annual, x='AÑO', y='MANIFIESTOS',
            text_auto='.2s', template='plotly_white',
            color_discrete_sequence=['#2563eb']
        )
        fig_envios.update_layout(xaxis_title="Año", yaxis_title="Manifiestos Sumados", margin=dict(l=10, r=10, t=30, b=10))
        st.plotly_chart(fig_envios, use_container_width=True)

    with col_d:
        st.markdown("##### 📈 Evolución Comparativa Peso vs Volumen")
        fig_peso_vol = px.line(
            df_annual, x='AÑO', y=['PESO', 'VOLUME'],
            markers=True, template='plotly_white',
            labels={'value': 'Cantidad Acumulada', 'variable': 'Métrica'},
            color_discrete_sequence=['#f59e0b', '#06b6d4']
        )
        fig_peso_vol.update_layout(xaxis_title="Año", yaxis_title="Total Acumulado", margin=dict(l=10, r=10, t=30, b=10))
        st.plotly_chart(fig_peso_vol, use_container_width=True)


# --- TAB 2: CONDUCTORES Y RUTAS ---
with tab2:
    st.subheader("Análisis de Conductores y Rutas")
    col_cond, col_rutas = st.columns(2)

    with col_cond:
        st.markdown("##### 🏆 Top 10 Conductores con Más Viajes")
        df_cond = df_filtered.groupby('NOMBRE_CONDUCTOR').size().reset_index(name='VIAJES')
        df_cond = df_cond[df_cond['NOMBRE_CONDUCTOR'] != 'N/A']
        df_top_cond = df_cond.sort_values(by='VIAJES', ascending=False).head(10)

        if not df_top_cond.empty:
            fig_cond = px.bar(
                df_top_cond, x='VIAJES', y='NOMBRE_CONDUCTOR', orientation='h',
                template='plotly_white', color_discrete_sequence=['#6366f1'],
                text_auto=True
            )
            fig_cond.update_layout(
                yaxis={'categoryorder': 'total ascending'},
                xaxis_title="Cantidad de Viajes Realizados", yaxis_title="",
                margin=dict(l=10, r=10, t=30, b=10)
            )
            st.plotly_chart(fig_cond, use_container_width=True)
        else:
            st.info("No hay información de conductores disponible en el archivo cargado.")

    with col_rutas:
        st.markdown("##### 📍 Top 10 Rutas de Mayor Impacto (Por Peso)")
        df_rutas = df_filtered.groupby('RUTA')['PESO'].sum().reset_index()
        df_top_rutas = df_rutas.sort_values(by='PESO', ascending=False).head(10)
        
        fig_rutas = px.bar(
            df_top_rutas, x='PESO', y='RUTA', orientation='h',
            template='plotly_white', color_discrete_sequence=['#06b6d4'],
            text_auto='.2s'
        )
        fig_rutas.update_layout(
            yaxis={'categoryorder': 'total ascending'},
            xaxis_title="Peso Total Movilizado (kg)", yaxis_title="",
            margin=dict(l=10, r=10, t=30, b=10)
        )
        st.plotly_chart(fig_rutas, use_container_width=True)


# --- TAB 3: OPERACIÓN Y EFICIENCIA ---
with tab3:
    st.subheader("Eficiencia Operativa y Terceros")
    col_op1, col_op2 = st.columns(2)

    with col_op1:
        st.markdown("##### 📉 Ocupación Promedio (%) por Año")
        fig_ocupacion = px.line(
            df_annual, x='AÑO', y='VOLUME',
            markers=True, template='plotly_white',
            color_discrete_sequence=['#ef4444']
        )
        fig_ocupacion.update_layout(xaxis_title="Año", yaxis_title="Promedio Capacidad (%)", margin=dict(l=10, r=10, t=30, b=10))
        st.plotly_chart(fig_ocupacion, use_container_width=True)

    with col_op2:
        st.markdown("##### ⏰ Distribución de Horarios de Despacho (Horas Pico)")
        df_hora = df_filtered.groupby('HORA_SIMPLE')['MANIFIESTOS'].sum().reset_index()
        fig_hora = px.bar(
            df_hora, x='HORA_SIMPLE', y='MANIFIESTOS',
            template='plotly_white', color_discrete_sequence=['#8b5cf6']
        )
        fig_hora.update_layout(xaxis_title="Hora del Día", yaxis_title="Despachos Totales", margin=dict(l=10, r=10, t=30, b=10))
        st.plotly_chart(fig_hora, use_container_width=True)

    st.markdown("##### 🤝 Participación por Operador Logístico")
    df_op = df_filtered.groupby('OPERADOR')['MANIFIESTOS'].sum().reset_index()
    df_op = df_op[df_op['MANIFIESTOS'] > 0]
    
    fig_operador = px.pie(
        df_op, names='OPERADOR', values='MANIFIESTOS',
        hole=0.45, template='plotly_white'
    )
    fig_operador.update_traces(textposition='inside', textinfo='percent+label')
    fig_operador.update_layout(showlegend=True, margin=dict(l=10, r=10, t=30, b=10))
    st.plotly_chart(fig_operador, use_container_width=True)


# --- TAB 4: GENERACIÓN DE REPORTES (PUNTO NO 5) ---
with tab4:
    st.subheader("📄 Centro de Reportabilidad Ejecutiva Automatizada")
    st.markdown("Genera e imprime reportes oficiales ajustados dinámicamente a los filtros seleccionados en la barra lateral.")

    rep_col1, rep_col2 = st.columns(2)

    with rep_col1:
        st.markdown("""
        <div class="report-card">
            <h4>📕 Informe Ejecutivo en PDF</h4>
            <p>Genera un documento PDF formal, listo para presentar a gerencia con:</p>
            <ul>
                <li>Resumen de KPIs estratégicos</li>
                <li>Tablas agregadas de flujo anual</li>
                <li>Top 5 Conductores y Rutas principales</li>
                <li>Hallazgos y alertas operativas</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        
        pdf_bytes = generate_pdf_report(df_filtered, df_annual)
        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M")
        
        st.download_button(
            label="📥 Descargar Informe Ejecutivo (PDF)",
            data=pdf_bytes,
            file_name=f"Informe_Ejecutivo_Logistica_{timestamp_str}.pdf",
            mime="application/pdf",
            use_container_width=True
        )

    with rep_col2:
        st.markdown("""
        <div class="report-card">
            <h4>📊 Libro de Trabajo en Excel Multi-Pestaña (.xlsx)</h4>
            <p>Exporta un libro de Excel estilizado profesionalmente con:</p>
            <ul>
                <li><b>Pestaña 1:</b> Tarjetas de KPIs y Resumen Anual</li>
                <li><b>Pestaña 2:</b> Registros detallados de la operación</li>
                <li>Colores corporativos, bordes y formatos numéricos</li>
                <li>Ajuste automático de columnas</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        
        excel_bytes = generate_excel_report(df_filtered, df_annual)
        
        st.download_button(
            label="📊 Descargar Reporte Multi-Pestaña (Excel .xlsx)",
            data=excel_bytes,
            file_name=f"Reporte_Logistico_Estilizado_{timestamp_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )


# --- TAB 5: DATOS AUDITABLES ---
with tab5:
    st.subheader("🔍 Tabla de Datos Auditables")
    
    csv_data = df_filtered.to_csv(index=False, sep=";").encode('latin1')
    st.download_button(
        label="📥 Descargar Reporte en CSV",
        data=csv_data,
        file_name="Reporte_Logistico_Filtrado.csv",
        mime="text/csv"
    )

    cols_a_mostrar = [c for c in ['AÑO', 'FECHA', 'HORA', 'RUTA', 'OPERADOR', 'NOMBRE_CONDUCTOR', 'PLACA', 'MANIFIESTOS', 'PESO', 'PIEZAS', 'CAPACIDAD'] if c in df_filtered.columns]
    st.dataframe(df_filtered[cols_a_mostrar], use_container_width=True)